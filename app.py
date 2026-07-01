"""
North Texas Sales Rep Sourcing — web app
Run-and-download tool. Users pick a location and size; verticals are fixed
(baked into the Apify task). Deploys free to Streamlit Community Cloud.
"""
import io, datetime, re, json, os, zipfile, urllib.parse
from email.message import EmailMessage
import pandas as pd
import streamlit as st
from apify_client import ApifyClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- config (token + task come from Streamlit secrets) ----
TASK_ID = st.secrets.get("TASK_ID", "natural_viburnum~north-texas-sales-reps-a")
# Dropdown options. Texas metros first (primary sourcing area), then every
# Albireo Energy office location from the global footprint (22 U.S. + 4 Europe).
LOCATION_PRESETS = [
    # Broad / no specific city
    "Remote",
    # Texas metros
    "Dallas-Fort Worth", "Dallas", "Fort Worth", "Austin", "Houston", "San Antonio", "Texas",
    # Albireo U.S. offices
    "Redmond, WA", "Vancouver, WA", "Anaheim, CA", "Poway, CA", "Tempe, AZ", "Tucson, AZ",
    "Salt Lake City, UT", "Denver, CO", "Omaha, NE", "Huntsville, AL", "Hoover, AL",
    "Montgomery, AL", "Tampa, FL", "Fort Lauderdale, FL", "Miami, FL", "Moosic, PA",
    "Sterling, VA", "New Castle, DE", "Gambrills, MD", "Edison, NJ", "New York City, NY",
    "Norwood, MA", "Chelmsford, MA",
    # Albireo Europe offices
    "London, United Kingdom", "Westerham, United Kingdom", "Scotland, United Kingdom",
    "Midlands, United Kingdom",
]

# Each location expands to its metropolitan area + nearby cities (~50 mi), sent to the
# search as an OR list so we cover the metro (and, if the metro name doesn't resolve,
# the surrounding cities). Map is tunable in location_areas.json.
@st.cache_data(show_spinner=False)
def load_location_areas():
    try:
        with open(os.path.join(os.path.dirname(__file__), "location_areas.json"), encoding="utf-8") as f:
            return {k: v for k, v in json.load(f).items() if not k.startswith("_")}
    except Exception:
        return {}

LOCATION_AREAS = load_location_areas()

def expand_locations(locs):
    """Expand chosen locations to their metro-area + nearby-city OR list. Unknown /
    custom locations pass through as-is ('Remote' -> nationwide United States)."""
    out = []
    for loc in locs:
        area = LOCATION_AREAS.get(loc)
        if area:
            out.extend(area)
        elif str(loc).strip().lower() == "remote":
            out.append("United States")
        else:
            out.append(loc)
    return list(dict.fromkeys(out))

# Fixed search configuration (your verticals). Location + maxItems are set per run
# from the UI below; everything else stays locked so targeting never drifts.
TITLES = ["Sales Rep", "PLC", "Data Centers", "Power Quality", "Building Automation",
          "Sales", "Account Executive", "Business Development", "Sales Engineer", "Territory Manager"]
BASE_INPUT = {
    "autoQuerySegmentation": False,
    "currentJobTitles": TITLES,
    "pastJobTitles": TITLES,
    "industryIds": ["147", "923", "453", "983", "2468", "382"],
    "profileLanguages": ["English"],
    "profileScraperMode": "Full + email search",
    "recentlyChangedJobs": False,
    "recentlyPostedOnLinkedIn": False,
}

# ===================== Role profiles + Fit Score =====================
# Roles, their keywords/industries/verticals, and per-role Fit Score weights are
# generated from AI_Sourcing_Albireo_Energy.xlsx into role_profiles.json (committed
# alongside this app, so it deploys to Streamlit Cloud and stays easy to tune).
@st.cache_data(show_spinner=False)
def load_role_data():
    with open(os.path.join(os.path.dirname(__file__), "role_profiles.json"), encoding="utf-8") as f:
        return json.load(f)

ROLE_DATA = load_role_data()
ROLES = {r["role"]: r for r in ROLE_DATA["roles"]}     # role name -> profile
ROLE_NAMES = [r["role"] for r in ROLE_DATA["roles"]]   # already grouped BAS then PLC
COMPETITORS = ROLE_DATA["competitors"]                 # category -> [company, ...]
# Flat list of competitor company names (slashes split) for "competitor experience".
_COMP_NAMES = []
for _cat_names in COMPETITORS.values():
    for _co in _cat_names:
        for _part in str(_co).split('/'):
            _p = _part.strip()
            if len(_p) >= 2:
                _COMP_NAMES.append(_p)

# Map a role's Fit-Score criterion name to a computable signal. Most role-specific
# "experience" criteria are captured by the role's skills/terms keyword match.
def _criterion_signal_key(name):
    n = (name or '').lower()
    if 'competitor' in n:                       return 'competitor'
    if 'geographic' in n or 'location' in n:    return 'location'
    if 'vertical' in n:                         return 'vertical'
    if 'industry' in n:                         return 'industry'
    if 'title' in n:                            return 'title'
    return 'skills'

_TITLE_SYNONYMS = {'sr': 'senior', 'jr': 'junior', 'vp': 'vice president', 'svp': 'senior vice president',
    'ae': 'account executive', 'bd': 'business development', 'bdr': 'business development',
    'mgr': 'manager', 'exec': 'executive', 'rep': 'representative', 'eng': 'engineer',
    'biz': 'business', 'dev': 'development'}

# ============ data helpers (same logic as the local pipeline) ============
def flatten(obj, prefix=''):
    out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            out.update(flatten(v, f"{prefix}/{k}" if prefix else k))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            out.update(flatten(v, f"{prefix}/{i}"))
    else:
        out[prefix] = obj
    return out

def val(d, key):
    v = d.get(key)
    if v is None:
        return ''
    try:
        if isinstance(v, float) and pd.isna(v):
            return ''
    except Exception:
        pass
    s = str(v).strip()
    return '' if s.lower() == 'nan' else s

VMAP = {
    'Data Center': ['data center','datacenter','critical facilit','mission critical','colocation'],
    'Building Automation': ['building automation','hvac control','bas ','automated logic','distech','niagara','building control'],
    'PLC / Industrial Automation': ['plc','programmable logic','scada','industrial automation'],
    'Energy Management': ['energy management','energy efficiency','sustainab'],
    'Power Systems / Power Quality': ['power quality','power systems','switchgear','uninterruptible',' ups ','electrical distribution','power distribution'],
    'HVAC / Heating & Air': ['hvac','heating and air','mechanical contractor','refrigeration'],
}
COMPANY_VERTICAL = {
    'Building Automation': ['johnson controls','honeywell','siemens','schneider','automated logic','distech','alerton','delta controls','resource data management','reliable controls','kmc controls'],
    'HVAC / Heating & Air': ['lennox','trane','carrier','daikin','texas airsystems','york','goodman','rheem','ruud','aaon','mitsubishi electric','airsystems'],
    'PLC / Industrial Automation': ['ifm','sensopart','vaisala','rockwell','emerson','abb','beckhoff','omron','banner engineering','pepperl','balluff','sick ','turck','wago'],
    'Power Systems / Power Quality': ['eaton','vertiv','hitachi energy','apc','resa power','generac','cummins','kohler','mitsubishi electric power','schweitzer'],
    'Electrical Distribution': ['graybar','rexel','consolidated electrical','border states','wesco','sonepar'],
}
def vmatch(d):
    company = val(d,'currentPosition/0/companyName').lower()
    blob = ' '.join([val(d,'headline'), val(d,'currentPosition/0/position'), company, val(d,'about')]).lower()
    tags = []
    for k, names in COMPANY_VERTICAL.items():
        if any(n in company for n in names):
            tags.append(k)
    for k, kw in VMAP.items():
        if k not in tags and any(w in blob for w in kw):
            tags.append(k)
    return ', '.join(tags) if tags else '—'

def prev_roles(d):
    cur = (val(d,'currentPosition/0/position').lower(), val(d,'currentPosition/0/companyName').lower())
    seen, out = set(), []
    for i in range(25):
        pos, comp, dur = val(d,f'experience/{i}/position'), val(d,f'experience/{i}/companyName'), val(d,f'experience/{i}/duration')
        if not pos and not comp:
            continue
        k = (pos.lower(), comp.lower())
        if k == cur or k in seen:
            continue
        seen.add(k)
        s = f"{pos} @ {comp}" if comp else pos
        if dur:
            s += f"  ({dur})"
        out.append('• ' + s)
    return '\n'.join(out[:4])

def skills(d):
    s = []
    for i in range(8):
        t = val(d, f'topSkills/{i}')
        if t:
            s.append(t)
    for i in range(60):
        n = val(d, f'skills/{i}/name')
        if n and n not in s:
            s.append(n)
    return ', '.join(s[:8])

def email(d):
    for k in ['email','workEmail','emailAddress','professionalEmail','personalEmail',
              'emails/0','emails/0/email','emails/0/value','emailSearch/email',
              'emailSearch/0/email','contactInfo/email','contact/email','emailAddresses/0']:
        e = val(d, k)
        if e and '@' in e:
            return e
    return ''

# ============================ Fit Score ============================
def _tokens(s):
    """Lower-cased word tokens with a few title abbreviations expanded."""
    s = re.sub(r'[^a-z0-9 ]', ' ', (s or '').lower())
    out = []
    for t in s.split():
        out.extend(_TITLE_SYNONYMS.get(t, t).split())
    return set(out)

def _title_match_fraction(title, targets):
    """Fuzzy/partial match of one title against the target title list -> 0..1.
    Full credit when every word of a target title appears in the candidate
    title (so 'Sr. Account Executive' fully matches 'Account Executive');
    partial credit for partial word overlap."""
    toks = _tokens(title)
    if not toks:
        return 0.0
    best = 0.0
    for tgt in targets:
        tt = _tokens(tgt)
        if not tt:
            continue
        if tt <= toks:
            return 1.0
        best = max(best, len(tt & toks) / len(tt))
    return best

def _former_titles(d):
    out = []
    for i in range(25):
        p = val(d, f'experience/{i}/position')
        if p:
            out.append(p)
    for i in range(1, 6):  # any current positions beyond the primary one
        p = val(d, f'currentPosition/{i}/position')
        if p:
            out.append(p)
    return out

def _industry_text(d):
    for k in ['currentPosition/0/companyIndustry', 'currentPosition/0/industry',
              'company/industry', 'companyIndustry', 'industry', 'industryName', 'occupation']:
        t = val(d, k)
        if t:
            return t.lower()
    return ''

def _cand_blob(d, rec):
    """Everything searchable about a candidate, lower-cased, for keyword matching."""
    parts = [rec.get('Headline', ''), rec.get('Current Title', ''), rec.get('Current Company', ''),
             rec.get('Top Skills', ''), rec.get('Previous Roles', ''), val(d, 'about')]
    for i in range(25):
        parts.append(val(d, f'experience/{i}/position'))
        parts.append(val(d, f'experience/{i}/companyName'))
    for i in range(60):
        parts.append(val(d, f'skills/{i}/name'))
    return ' '.join(p for p in parts if p).lower()

def _kw_hit(kw, blob):
    """True if a keyword/phrase appears in the blob. Short or all-caps keywords
    (BAS, PLC, DDC, N2…) require word boundaries to avoid false positives."""
    k = (kw or '').strip()
    if not k:
        return False
    kl = k.lower()
    if len(kl) <= 4 or k.isupper():
        return re.search(r'(?<![a-z0-9])' + re.escape(kl) + r'(?![a-z0-9])', blob) is not None
    return kl in blob

def _kw_signal(keywords, blob, target):
    """Fraction in 0..1: number of distinct keyword hits, saturating at `target`."""
    hits = sum(1 for kw in keywords if _kw_hit(kw, blob))
    return min(1.0, hits / float(target)) if target else 0.0

def _title_signal(d, rec, job_titles):
    cur = _title_match_fraction(rec.get('Current Title', ''), job_titles)
    former = max((_title_match_fraction(t, job_titles) for t in _former_titles(d)), default=0.0)
    return max(cur, 0.7 * former)   # current title counts more than a past one

def _competitor_info(d, rec):
    """Return (matched competitor name, signal). Current employer match = 1.0,
    former employer match = 0.6, otherwise ('', 0.0)."""
    def hit(company):
        c = (company or '').strip().lower()
        if not c:
            return ''
        for name in _COMP_NAMES:
            nl = name.lower()
            if len(nl) <= 4:
                if re.search(r'(?<![a-z0-9])' + re.escape(nl) + r'(?![a-z0-9])', c):
                    return name
            elif nl in c or c in nl:
                return name
        return ''
    m = hit(rec.get('Current Company', ''))
    if m:
        return m, 1.0
    for i in range(25):
        m = hit(val(d, f'experience/{i}/companyName'))
        if m:
            return m, 0.6
    return '', 0.0

# Location scoring is relative to the run's target location (no hardcoded city).
_STATE_ABBR = {'al','ak','az','ar','ca','co','ct','de','fl','ga','hi','id','il','in','ia','ks',
    'ky','la','me','md','ma','mi','mn','ms','mo','mt','ne','nv','nh','nj','nm','ny','nc','nd',
    'oh','ok','or','pa','ri','sc','sd','tn','tx','ut','vt','va','wa','wv','wi','wy'}
_STATE_NAME_ABBR = {'alabama':'al','alaska':'ak','arizona':'az','arkansas':'ar','california':'ca',
    'colorado':'co','connecticut':'ct','delaware':'de','florida':'fl','georgia':'ga','hawaii':'hi',
    'idaho':'id','illinois':'il','indiana':'in','iowa':'ia','kansas':'ks','kentucky':'ky',
    'louisiana':'la','maine':'me','maryland':'md','massachusetts':'ma','michigan':'mi',
    'minnesota':'mn','mississippi':'ms','missouri':'mo','montana':'mt','nebraska':'ne',
    'nevada':'nv','ohio':'oh','oklahoma':'ok','oregon':'or','pennsylvania':'pa','tennessee':'tn',
    'texas':'tx','utah':'ut','vermont':'vt','virginia':'va','washington':'wa','wisconsin':'wi',
    'wyoming':'wy'}
_LOC_STOP = {'area','greater','metro','metropolitan','city','of','the','region','county'}

def _loc_tokens(s):
    s = re.sub(r'[^a-z0-9 ]', ' ', (s or '').lower())
    out = []
    for t in s.split():
        if not t or t in _LOC_STOP:
            continue
        out.append(_STATE_NAME_ABBR.get(t, t))  # normalize full state names -> abbr
    return out

def _location_fraction(cand_loc, target_loc):
    """Higher when the candidate is in/near the run's target location.
    Returns None when there's no location context (so it's left out of the
    score rather than penalizing anyone)."""
    if not (target_loc or '').strip():
        return None
    cand = set(_loc_tokens(cand_loc))
    if not cand:
        return 0.0
    tgt = _loc_tokens(target_loc)
    cities = [t for t in tgt if t not in _STATE_ABBR]
    states = [t for t in tgt if t in _STATE_ABBR]
    if cities:
        if any(c in cand for c in cities):
            return 1.0                      # same city / metro / region
        if any(s in cand for s in states):
            return 0.5                      # same state -> "near"
        return 0.0
    # state-only target (e.g. "Texas") — being in that state is a full match
    return 1.0 if any(s in cand for s in states) else 0.0

def _location_fraction_multi(cand_loc, targets):
    """Best location match across the selected target locations (0..1). Remote /
    nationwide targets don't contribute; if ALL targets are remote, returns None
    so location is dropped from the score (nobody penalized)."""
    specifics = [t for t in (targets or [])
                 if str(t).strip().lower() not in ("remote", "united states", "")]
    if not specifics:
        return None
    best = 0.0
    for t in specifics:
        f = _location_fraction(cand_loc, t)
        if f is not None:
            best = max(best, f)
    return best

def role_fit_score(d, rec, role, target_locations, comp_signal):
    """0–100 fit score using the selected role's own criteria + weights. Each
    criterion maps to a computable signal (title / skills-terms / industry /
    vertical / competitor experience / location). The role-specific 'experience'
    criteria are captured by the role's skills/terms keyword match. Score is
    normalized over whichever criteria apply (location drops out with no run
    location), so it always reads 0–100."""
    blob = _cand_blob(d, rec)
    signals = {
        'title':      _title_signal(d, rec, role.get('job_titles', [])),
        'skills':     _kw_signal(role.get('skills', []), blob, target=5),
        'industry':   _kw_signal(role.get('industries', []), _industry_text(d) + ' ' + blob, target=2),
        'vertical':   _kw_signal(role.get('verticals', []), blob, target=2),
        'competitor': comp_signal,
        'location':   _location_fraction_multi(rec.get('Location', ''), target_locations),
    }
    acc = tot = 0.0
    for crit in role.get('criteria', []):
        s = signals.get(_criterion_signal_key(crit['name']))
        if s is None:          # location criterion but the run has no location -> drop its weight
            continue
        acc += crit['weight'] * s
        tot += crit['weight']
    return round(acc / tot * 100) if tot else 0

# Display columns: Fit Score first, then Location after Current Title, then a
# Competitor flag right after Current Company.
COLUMNS = ['Fit Score','First Name','Last Name','Current Title','Location','Current Company',
           'Competitor','Vertical Match','Headline','Tenure','Open To Work','Email','Previous Roles',
           'Top Skills','LinkedIn URL','Company LinkedIn','Connections','Summary']

def condense(d):
    otw = val(d,'openToWork').lower()
    otw = 'Yes' if otw == 'true' else ('No' if otw == 'false' else '')
    return {
        'First Name': val(d,'firstName'), 'Last Name': val(d,'lastName'),
        'Current Title': val(d,'currentPosition/0/position'),
        'Current Company': val(d,'currentPosition/0/companyName'),
        'Vertical Match': vmatch(d), 'Headline': val(d,'headline'),
        'Location': val(d,'location/linkedinText'),
        'Tenure': val(d,'currentPosition/0/duration'), 'Open To Work': otw,
        'Email': email(d), 'Previous Roles': prev_roles(d), 'Top Skills': skills(d),
        'LinkedIn URL': val(d,'linkedinUrl'),
        'Company LinkedIn': val(d,'currentPosition/0/companyLinkedinUrl'),
        'Connections': val(d,'connectionsCount'), 'Summary': val(d,'about'),
    }

def _excel_summary_sheet(ws, summary):
    """Write a small run-summary block (counts) onto its own sheet."""
    title_fill = PatternFill('solid', fgColor='1F4E5F')
    head = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    label = Font(name='Calibri', bold=True, size=11); base = Font(name='Calibri', size=11)
    ws['A1'] = 'Run Summary'; ws['A1'].font = head; ws['A1'].fill = title_fill
    ws['B1'].fill = title_fill
    rows = [
        ('Role', summary.get('Role', '')),
        ('Search location', summary.get('Location', '')),
        ('Date', summary.get('Timestamp', '')),
        ('Profiles requested', summary.get('Requested', '')),
        ('Profiles found', summary.get('Found', '')),
        ('New (added to master)', summary.get('New', '')),
        ('Duplicates (already on file)', summary.get('Dupes', '')),
        ('Master list total', summary.get('MasterTotal', '')),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(i, 1, k).font = label
        ws.cell(i, 2, v).font = base
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 28
    ws.sheet_view.showGridLines = False

def build_excel_bytes(rows, cols=COLUMNS, summary=None):
    wb = Workbook()
    if summary is not None:
        ws_sum = wb.active; ws_sum.title = 'Summary'; _excel_summary_sheet(ws_sum, summary)
        ws = wb.create_sheet('Sales Reps')
    else:
        ws = wb.active; ws.title = 'Sales Reps'
    hf = PatternFill('solid', fgColor='1F4E5F'); hfont = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    band = PatternFill('solid', fgColor='F2F6F7'); thin = Side(style='thin', color='E2E2E2')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    lf = Font(name='Calibri', color='0563C1', underline='single', size=11); base = Font(name='Calibri', size=11)
    wrap = {'Current Title','Vertical Match','Headline','Previous Roles','Top Skills','Location'}
    links = {'LinkedIn URL','Company LinkedIn'}
    ws.append(cols)
    for ci, h in enumerate(cols, 1):
        c = ws.cell(1, ci); c.fill = hf; c.font = hfont; c.border = bd
        c.alignment = Alignment(vertical='center', wrap_text=True)
    for ri, row in enumerate(rows, start=2):
        for ci, h in enumerate(cols, 1):
            cell = ws.cell(ri, ci, row.get(h, '')); cell.border = bd
            cell.alignment = Alignment(vertical='top', horizontal=('right' if h in ('Connections','Fit Score') else 'left'), wrap_text=(h in wrap))
            if h in links and cell.value:
                cell.hyperlink = cell.value; cell.font = lf
            else:
                cell.font = base
            if ri % 2 == 0:
                cell.fill = band
    W = {'Fit Score':10,'Date Added':13,'Sourced Role':26,'Search Location':18,'First Name':12,
         'Last Name':13,'Current Title':26,'Current Company':22,'Competitor':20,'Vertical Match':22,
         'Headline':38,'Location':24,'Tenure':12,'Open To Work':11,'Email':30,'Previous Roles':46,
         'Top Skills':28,'LinkedIn URL':34,'Company LinkedIn':34,'Connections':12,'Summary':55}
    for ci, h in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = W.get(h, 16)
    ws.freeze_panes = 'A2'; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.sheet_view.showGridLines = False
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def _dsid(r):
    if isinstance(r, dict):
        return r.get("defaultDatasetId") or r.get("default_dataset_id")
    return getattr(r, "default_dataset_id", None) or getattr(r, "defaultDatasetId", None)

# ============ persistence: Google Sheets master list + search history ============
# Keeps a running master list (deduped by LinkedIn URL) and a log of every search,
# so new-vs-duplicate counts work across runs and days. If the Google Sheet secrets
# aren't configured yet, everything falls back to in-session memory so the app still
# runs — see SETUP_GOOGLE_SHEETS.md to connect a sheet.
# Master sheet order: Date Added, Fit Score (2nd), the role it was sourced for, then the rest.
MASTER_HEADERS = ["Date Added", "Fit Score", "Sourced Role", "Search Location"] + [c for c in COLUMNS if c != "Fit Score"]
SEARCH_HEADERS = ["Timestamp", "Role", "Location", "Requested", "Found", "New", "Dupes"]

def _norm_url(u):
    u = (u or "").strip().lower()
    for p in ("https://", "http://", "www."):
        u = u.replace(p, "")
    return u.rstrip("/")

def gs_configured():
    try:
        has_creds = ("gcp_service_account_json" in st.secrets
                     or "gcp_service_account" in st.secrets)
        return has_creds and "MASTER_SHEET_ID" in st.secrets
    except Exception:
        return False

@st.cache_resource(show_spinner=False)
def _gs_book():
    import json, gspread
    from google.oauth2.service_account import Credentials
    # Accept either the whole service-account JSON pasted as one string
    # (easiest), or a classic [gcp_service_account] TOML table.
    if "gcp_service_account_json" in st.secrets:
        info = json.loads(st.secrets["gcp_service_account_json"])
    else:
        info = dict(st.secrets["gcp_service_account"])
    creds = Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/spreadsheets"])
    return gspread.authorize(creds).open_by_key(st.secrets["MASTER_SHEET_ID"])

# ---- cosmetic formatting for the Google Sheet tabs (best-effort) ----
_NAVY_RGB = {"red": 0.055, "green": 0.176, "blue": 0.322}  # #0E2D52
_COLW = {
    'Fit Score': 75, 'Date Added': 95, 'Sourced Role': 200, 'Search Location': 135, 'Role': 200,
    'First Name': 95, 'Last Name': 100, 'Competitor': 150,
    'Current Title': 190, 'Current Company': 165, 'Vertical Match': 165, 'Headline': 290,
    'Location': 175, 'Tenure': 95, 'Open To Work': 95, 'Email': 220, 'Previous Roles': 320,
    'Top Skills': 200, 'LinkedIn URL': 240, 'Company LinkedIn': 240, 'Connections': 100,
    'Summary': 380, 'Timestamp': 135, 'Requested': 95, 'Found': 80, 'New': 75, 'Dupes': 80,
}

def _format_worksheet(ws, headers):
    """Frozen bold header, sane column widths, zebra rows. Cosmetic only —
    each piece is best-effort so styling can never break the app."""
    header = ws.row_values(1) or list(headers)  # style the sheet's actual columns
    ncols = len(header)
    last = get_column_letter(ncols)
    try:
        ws.freeze(rows=1)
        ws.format(f"A1:{last}1", {
            "backgroundColor": _NAVY_RGB, "verticalAlignment": "MIDDLE",
            "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1},
                           "bold": True, "fontSize": 10}})
    except Exception:
        pass
    try:
        ws.spreadsheet.batch_update({"requests": [
            {"updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS",
                          "startIndex": i, "endIndex": i + 1},
                "properties": {"pixelSize": _COLW.get(h, 130)}, "fields": "pixelSize"}}
            for i, h in enumerate(header)]})
    except Exception:
        pass
    try:  # alternating row shading; ignored if a band already exists
        ws.spreadsheet.batch_update({"requests": [{"addBanding": {"bandedRange": {
            "range": {"sheetId": ws.id, "startRowIndex": 0,
                      "startColumnIndex": 0, "endColumnIndex": ncols},
            "rowProperties": {"headerColor": _NAVY_RGB,
                              "firstBandColor": {"red": 1, "green": 1, "blue": 1},
                              "secondBandColor": {"red": 0.949, "green": 0.965, "blue": 0.969}}}}}]})
    except Exception:
        pass

def _maybe_format(ws, title, headers, force=False):
    done = st.session_state.setdefault("_ws_formatted", set())
    if force or title not in done:
        _format_worksheet(ws, headers)
        done.add(title)

def _ws(title, headers):
    import gspread
    sh = _gs_book()
    created = False
    try:
        ws = sh.worksheet(title)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=title, rows=2000, cols=len(headers))
        ws.append_row(headers, value_input_option="USER_ENTERED")
        created = True
    if not ws.row_values(1):
        ws.append_row(headers, value_input_option="USER_ENTERED")
    _maybe_format(ws, title, headers, force=created)
    return ws

# Cache sheet reads so we don't hit Google's per-minute read quota on every rerun.
# Cleared explicitly after any write so the UI still reflects changes immediately.
@st.cache_data(ttl=120, show_spinner=False)
def _read_records(title):
    hdrs = {"Master": MASTER_HEADERS, "Searches": SEARCH_HEADERS, "Shortlist": SHORTLIST_HEADERS}
    return _ws(title, hdrs[title]).get_all_records()

def _invalidate_reads():
    try:
        _read_records.clear()
    except Exception:
        pass

def load_master():
    """Return list of dicts (oldest first) for every unique rep saved so far."""
    if gs_configured():
        try:
            return _read_records("Master")
        except Exception as e:
            st.warning(f"Couldn't read the master sheet: {e}")
            return []
    return st.session_state.setdefault("master", [])

def load_searches():
    """Return list of dicts (oldest first) for every search run so far."""
    if gs_configured():
        try:
            return _read_records("Searches")
        except Exception as e:
            st.warning(f"Couldn't read search history: {e}")
            return []
    return st.session_state.setdefault("searches", [])

def _ensure_columns(ws, needed):
    """Make sure every needed column exists in the sheet's header, appending any
    new ones (e.g. 'Fit Score') at the end so existing rows stay aligned.
    Returns the sheet's actual header order to write rows against."""
    header = ws.row_values(1)
    if not header:
        ws.update(range_name="A1", values=[list(needed)])
        return list(needed)
    missing = [h for h in needed if h not in header]
    if missing:
        header = header + missing
        ws.update(range_name="A1", values=[header])
    return header

def save_master(rows):
    if not rows:
        return
    if gs_configured():
        ws = _ws("Master", MASTER_HEADERS)
        header = _ensure_columns(ws, MASTER_HEADERS)   # adds Fit Score column if absent
        ws.append_rows([[r.get(h, "") for h in header] for r in rows],
                       value_input_option="USER_ENTERED")
        _invalidate_reads()
    else:
        st.session_state.setdefault("master", []).extend(rows)

def save_search(rec):
    if gs_configured():
        ws = _ws("Searches", SEARCH_HEADERS)
        header = _ensure_columns(ws, SEARCH_HEADERS)   # adds Role column if absent
        ws.append_row([rec.get(h, "") for h in header],
                      value_input_option="USER_ENTERED")
        _invalidate_reads()
    else:
        st.session_state.setdefault("searches", []).append(rec)

# ============================ shortlist ============================
SHORTLIST_HEADERS = ["Added", "Sourced Role", "Fit Score", "First Name", "Last Name", "Email",
                     "Current Title", "Current Company", "Location", "LinkedIn URL", "Status"]

def load_shortlist():
    if gs_configured():
        try:
            return _read_records("Shortlist")
        except Exception as e:
            st.warning(f"Couldn't read the shortlist: {e}")
            return []
    return st.session_state.setdefault("shortlist", [])

def add_to_shortlist(rows):
    """Append people not already on the shortlist (deduped by LinkedIn URL)."""
    existing = {_norm_url(r.get("LinkedIn URL", "")) for r in load_shortlist()}
    fresh = [r for r in rows if _norm_url(r.get("LinkedIn URL", "")) not in existing]
    if not fresh:
        return 0
    if gs_configured():
        ws = _ws("Shortlist", SHORTLIST_HEADERS)
        header = _ensure_columns(ws, SHORTLIST_HEADERS)
        ws.append_rows([[r.get(h, "") for h in header] for r in fresh],
                       value_input_option="USER_ENTERED")
        _invalidate_reads()
    else:
        st.session_state.setdefault("shortlist", []).extend(fresh)
    return len(fresh)

def replace_shortlist(rows):
    """Overwrite the whole shortlist (used for removals)."""
    if gs_configured():
        ws = _ws("Shortlist", SHORTLIST_HEADERS)
        ws.clear()
        ws.update(range_name="A1",
                  values=[SHORTLIST_HEADERS] + [[r.get(h, "") for h in SHORTLIST_HEADERS] for r in rows])
        _invalidate_reads()
    else:
        st.session_state["shortlist"] = list(rows)

# ============================ outreach email (drafts only — no sending) ============================
DEFAULT_SUBJECT = "Exploring a {role} opportunity at Albireo Energy"
DEFAULT_BODY = (
    "Hi {first_name},\n\n"
    "I came across your background as {title} at {company} and was really impressed. "
    "At Albireo Energy we're growing our team, and I think you could be a strong fit for a "
    "{role} position we're hiring for.\n\n"
    "Would you be open to a quick conversation this week?\n\n"
    "Best regards,\n"
    "[Your name]\n"
    "Albireo Energy"
)

def fill_template(text, rec):
    repl = {
        "{first_name}": rec.get("First Name", ""), "{last_name}": rec.get("Last Name", ""),
        "{role}": rec.get("Sourced Role", "") or rec.get("Role", ""),
        "{title}": rec.get("Current Title", ""), "{company}": rec.get("Current Company", ""),
    }
    out = text or ""
    for k, v in repl.items():
        out = out.replace(k, str(v))
    return out

def build_eml(to, subject, body, sender):
    m = EmailMessage()
    if sender:
        m["From"] = sender
    m["To"] = to or ""
    m["Subject"] = subject
    m.set_content(body)
    return m.as_bytes()

def build_drafts_zip(people, subject_tmpl, body_tmpl, sender):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for i, r in enumerate(people, 1):
            name = f"{i:02d}_{r.get('First Name','')}_{r.get('Last Name','')}".strip().replace(' ', '_')
            z.writestr(f"{name or 'draft'}.eml",
                       build_eml(r.get("Email", ""), fill_template(subject_tmpl, r),
                                 fill_template(body_tmpl, r), sender))
    buf.seek(0)
    return buf.getvalue()

def build_drafts_csv(people, subject_tmpl, body_tmpl):
    rows = [{"To": r.get("Email", ""), "Subject": fill_template(subject_tmpl, r),
             "Body": fill_template(body_tmpl, r)} for r in people]
    return pd.DataFrame(rows, columns=["To", "Subject", "Body"]).to_csv(index=False).encode("utf-8")

def outlook_compose_url(to, subject, body):
    """Deep link that opens a pre-filled compose window in the user's Outlook."""
    qs = urllib.parse.urlencode({"to": to or "", "subject": subject, "body": body},
                                quote_via=urllib.parse.quote)
    return "https://outlook.office.com/mail/deeplink/compose?" + qs

# ============================ login (Microsoft SSO) ============================
# Only staff on this domain may use the app. Set to "" to allow any Microsoft account.
ALLOWED_EMAIL_DOMAIN = "albireoenergy.com"

def _auth_configured():
    try:
        return "auth" in st.secrets
    except Exception:
        return False

def _user_email():
    try:
        u = st.user
        for k in ("email", "preferred_username", "upn", "unique_name"):
            v = u.get(k) if hasattr(u, "get") else getattr(u, k, None)
            if v:
                return str(v).lower()
    except Exception:
        pass
    return ""

def require_login():
    """Gate the app behind Microsoft SSO once the [auth] secrets are configured.
    Until then it's a no-op, so the app keeps working during setup."""
    if not _auth_configured():
        return
    if not st.user.is_logged_in:
        st.markdown("### 🔐 Sign in required")
        st.write("This tool is restricted to Albireo Energy staff. Please sign in to continue.")
        st.button("Sign in with Microsoft", type="primary", on_click=st.login)
        st.stop()
    email = _user_email()
    if ALLOWED_EMAIL_DOMAIN and not email.endswith("@" + ALLOWED_EMAIL_DOMAIN):
        st.error(f"Access is restricted to @{ALLOWED_EMAIL_DOMAIN} accounts.")
        st.caption(f"Detected sign-in: {email or '(no email claim returned)'}")
        try:
            st.caption("Claims returned: " + ", ".join(sorted(dict(st.user).keys())))
        except Exception:
            pass
        st.button("Sign out", on_click=st.logout)
        st.stop()

# ============================ UI ============================
st.set_page_config(page_title="Albireo Energy · Sales Rep Sourcing", page_icon="🎯", layout="wide")

NAVY = "#0E2D52"; NAVY_DARK = "#081C36"; BLUE = "#1F6FB2"
GOLD = "#FDB813"; GOLD_DARK = "#E0A200"; INK = "#16263B"
GRAY = "#E5EAF0"; CARD = "#FFFFFF"; BORDER = "#D6DEE8"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
html, body, .stApp, button, input, textarea, select,
[data-testid="stMarkdownContainer"] {{ font-family: 'Inter', -apple-system, sans-serif; }}

.stApp {{ background: {GRAY}; }}
.block-container {{ padding-top: 2.2rem; padding-bottom: 2rem;
    padding-left: 3rem; padding-right: 3rem; max-width: 1320px; }}

/* header banner */
.ae-header {{
    background: linear-gradient(135deg, {NAVY} 0%, {NAVY_DARK} 100%);
    border-bottom: 5px solid {GOLD};
    border-radius: 18px; padding: 40px 44px; margin-bottom: 30px;
    box-shadow: 0 10px 30px rgba(8,28,54,0.25);
}}
.ae-eyebrow {{ color: {GOLD} !important; font-weight: 800; letter-spacing: 3px;
    font-size: 14px; text-transform: uppercase; margin: 0; }}
.ae-title {{ color: #ffffff !important; font-size: 46px; font-weight: 800;
    margin: 8px 0 0 0; line-height: 1.05; }}
.ae-chips {{ display: flex; flex-wrap: wrap; gap: 8px; margin-top: 18px; }}
.ae-chip {{ background: rgba(255,255,255,0.10); color: #cfdbea;
    border: 1px solid rgba(255,255,255,0.18); font-size: 12px; font-weight: 600;
    padding: 5px 13px; border-radius: 999px; }}

/* input section labels */
.ae-label {{ font-size: 22px; font-weight: 800; color: {NAVY}; margin: 2px 0 2px 0; }}
.ae-help  {{ font-size: 13px; color: #66788c; margin: 0 0 12px 0; }}

/* gold run button (navy text) */
div.stButton, div[data-testid="stButton"] {{ width: 100% !important; }}
div.stButton > button, div[data-testid="stButton"] > button {{
    background: linear-gradient(135deg, #FFC62E 0%, {GOLD} 100%);
    color: {NAVY}; font-size: 24px; font-weight: 800; letter-spacing: .4px;
    padding: 20px 0; border: none; border-radius: 14px; width: 100% !important;
    white-space: nowrap; line-height: 1.2;
    box-shadow: 0 8px 22px rgba(253,184,19,0.45); transition: all .15s ease;
}}
div.stButton > button:hover, div[data-testid="stButton"] > button:hover {{
    background: linear-gradient(135deg, {GOLD} 0%, {GOLD_DARK} 100%);
    color: {NAVY}; transform: translateY(-2px); box-shadow: 0 12px 28px rgba(253,184,19,0.52);
}}
div.stButton > button:active, div[data-testid="stButton"] > button:active {{ transform: translateY(0); }}

/* navy download button */
div.stDownloadButton > button {{
    background: {NAVY}; color: #fff; font-size: 15px; font-weight: 700;
    padding: 12px 26px; border: none; border-radius: 10px;
    box-shadow: 0 4px 14px rgba(14,45,82,0.25); transition: all .15s ease;
}}
div.stDownloadButton > button:hover {{ background: {NAVY_DARK}; color: #fff; transform: translateY(-1px); }}

/* input cards */
div[data-testid="stVerticalBlockBorderWrapper"] {{
    background: {CARD}; border-radius: 16px;
    box-shadow: 0 4px 18px rgba(14,45,82,0.09); border: 1px solid {BORDER} !important;
}}
div[data-testid="stVerticalBlockBorderWrapper"] > div {{ padding: 18px 22px; min-height: 150px; }}

/* empty state */
.ae-empty {{ background: #fff; border: 2px dashed #c4d0de; border-radius: 16px;
    padding: 60px 24px; text-align: center; margin-top: 8px; }}
.ae-empty .big {{ font-size: 46px; line-height: 1; }}
.ae-empty .t {{ font-size: 18px; font-weight: 700; color: {NAVY}; margin-top: 12px; }}
.ae-empty .s {{ font-size: 14px; color: #7a8a9c; margin-top: 4px; }}

/* branded results banner */
.ae-result {{ background: {BLUE}; color: #fff; font-weight: 700; font-size: 16px;
    padding: 13px 20px; border-radius: 12px; margin: 8px 0 16px 0;
    box-shadow: 0 4px 14px rgba(31,111,178,0.25); }}

/* top tabs */
div[data-testid="stTabs"] div[role="tablist"] {{
    gap: 4px; border-bottom: 2px solid {BORDER}; margin-bottom: 18px; }}
div[data-testid="stTabs"] button[role="tab"] {{
    padding: 10px 22px; border-radius: 10px 10px 0 0; }}
div[data-testid="stTabs"] button[role="tab"] p {{
    font-size: 16px; font-weight: 700; color: #66788c; }}
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {{
    background: #fff; }}
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] p {{
    color: {NAVY}; font-weight: 800; }}

/* footer */
.ae-footer {{ text-align: center; color: #8a98a8; font-size: 12px;
    margin-top: 36px; padding-top: 18px; border-top: 1px solid {BORDER}; }}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="ae-header">
    <p class="ae-eyebrow">Albireo Energy</p>
    <h1 class="ae-title">Sales Rep Sourcing</h1>
</div>
""", unsafe_allow_html=True)

require_login()  # Microsoft SSO gate (active once [auth] secrets are configured)
if _auth_configured() and st.user.is_logged_in:
    _lc = st.columns([6, 1])
    _lc[0].caption(f"🔐 Signed in as {_user_email()}")
    _lc[1].button("Log out", on_click=st.logout, use_container_width=True)

tab_search, tab_history, tab_master, tab_shortlist = st.tabs(
    ["🔎  New Search", "🕘  Previous Searches", "📒  Master List", "⭐  Shortlist"])

# ------------------------------- New Search -------------------------------
with tab_search:
    with st.container(border=True):
        st.markdown('<p class="ae-label">🧭 Role</p>', unsafe_allow_html=True)
        st.markdown('<p class="ae-help">What role are you sourcing for? Sets the search keywords and the Fit Score weights.</p>',
                    unsafe_allow_html=True)
        role_name = st.selectbox("Role", ROLE_NAMES, label_visibility="collapsed")
        _role = ROLES[role_name]
        with st.expander("What this role targets"):
            st.markdown(f"**Job titles searched:** {', '.join(_role['job_titles'])}")
            st.markdown(f"**Skills / terms:** {', '.join(_role['skills'][:18])}"
                        + ("…" if len(_role['skills']) > 18 else ""))
            st.markdown(f"**Industries:** {', '.join(_role['industries'])}")
            _ind_labels = ROLE_DATA.get("industry_labels", {})
            _filtered = [_ind_labels.get(i, i) for i in _role.get("industry_ids", [])]
            if _filtered:
                st.markdown(f"**LinkedIn industry filter:** {', '.join(_filtered)}")
            st.markdown(f"**Verticals:** {', '.join(_role['verticals'])}")
            st.markdown("**Fit Score weighting:** "
                        + " · ".join(f"{c['name']} {round(c['weight']*100)}%" for c in _role['criteria']))

    col1, col2 = st.columns(2, gap="large")
    with col1:
        with st.container(border=True):
            st.markdown('<p class="ae-label">📍 Location(s)</p>', unsafe_allow_html=True)
            st.markdown('<p class="ae-help">Pick one or more — e.g. Dallas, Fort Worth, Remote</p>',
                        unsafe_allow_html=True)
            sel_locs = st.multiselect("Location", LOCATION_PRESETS, default=["Dallas-Fort Worth"],
                                      label_visibility="collapsed", placeholder="Choose one or more locations…")
            custom_locs = st.text_input("Other locations", label_visibility="collapsed",
                                        placeholder="Add others, separated by ;  (e.g. Denver, CO; Phoenix, AZ)")
    with col2:
        with st.container(border=True):
            st.markdown('<p class="ae-label">🎯 Number of Profiles</p>', unsafe_allow_html=True)
            st.markdown('<p class="ae-help">How many matching reps you want this run</p>', unsafe_allow_html=True)
            size = st.slider("Number of profiles to pull", 10, 200, 25, step=5, label_visibility="collapsed")

    # combined, de-duplicated list of chosen locations (presets + custom)
    selected_locations = list(dict.fromkeys(
        list(sel_locs) + [x.strip() for x in custom_locs.split(';') if x.strip()]))

    otw_only = st.checkbox('🟢 Only show people flagged "Open to Work"')
    if any(s.strip().lower() != "remote" for s in selected_locations):
        st.caption("📍 Each location also covers its surrounding metro area / ~50-mile radius.")
    if any(s.strip().lower() == "remote" for s in selected_locations):
        st.caption("🌐 Remote = nationwide (United States) search; it won't penalize anyone by location.")

    st.write("")
    bcol = st.columns([1, 6, 1])
    with bcol[1]:
        clicked = st.button("🔎  Run Sourcing", type="primary", use_container_width=True)

    if not gs_configured():
        st.info("⚠️ Google Sheet not connected yet — searches still run, but the master list and "
                "history won't be saved between sessions. See **SETUP_GOOGLE_SHEETS.md** to connect one.")

    if clicked:
        st.session_state['last_run'] = None
        if not selected_locations:
            st.warning("Please choose at least one location.")
            st.stop()
        try:
            role = ROLES[role_name]
            location_label = ", ".join(selected_locations)
            # "Remote" -> nationwide (United States); a run that's ONLY remote drops
            # location from scoring, otherwise we score by the best-matching city.
            specifics = [s for s in selected_locations if s.strip().lower() != "remote"]
            remote_only = not specifics
            # expand each pick to its metro area + nearby cities (~50 mi) for coverage
            search_locations = expand_locations(selected_locations)[:40]
            with st.spinner(f"Searching LinkedIn for {role_name}… this usually takes a minute or two."):
                client = ApifyClient(st.secrets["APIFY_TOKEN"])
                # Target the search with the selected role's job titles + industries.
                run_input = dict(BASE_INPUT)
                run_input["currentJobTitles"] = role["job_titles"]
                run_input["pastJobTitles"] = role["job_titles"]
                run_input["industryIds"] = role.get("industry_ids") or BASE_INPUT["industryIds"]
                run_input["locations"] = search_locations
                # Backfill (conservative, up to ~2x): over-pull so that after dropping
                # profiles with no URL — and non-open-to-work people when that filter is
                # on — we can still land on the requested number of matching results.
                overpull = 2.0 if otw_only else 1.2
                run_input["maxItems"] = min(int(round(size * overpull)), 400)
                run = client.task(TASK_ID).call(task_input=run_input)
                run_status = run.get("status") if isinstance(run, dict) else getattr(run, "status", None)
                run_id = run.get("id") if isinstance(run, dict) else getattr(run, "id", None)
                ds = _dsid(run)
                items = [flatten(i) for i in client.dataset(ds).iterate_items()]
            recs = []
            for d in items:
                if not val(d, 'linkedinUrl'):
                    continue
                rec = condense(d)
                comp_name, comp_sig = _competitor_info(d, rec)
                rec['Competitor'] = comp_name
                # Fit Score uses the selected role's own criteria + weights; the expanded
                # location list gives credit to anyone in the metro / ~50-mile radius.
                rec['Fit Score'] = role_fit_score(d, rec, role, search_locations, comp_sig)
                recs.append(rec)
            if not recs:
                st.error(f"No usable profiles came back for **{role_name}** in **{location_label}**.")
                st.markdown(
                    f"- Apify run status: **{run_status}**\n"
                    f"- Raw items returned by the actor: **{len(items)}**\n"
                    f"- Of those, profiles with a LinkedIn URL: **{len(recs)}**\n"
                    f"- Run id: `{run_id}`")
                if len(items) > 0:
                    st.warning("The actor returned items but none had a LinkedIn URL — the profile "
                               "field names may have changed. Send me the run id above and I'll map it.")
                elif run_status and run_status != "SUCCEEDED":
                    st.warning(f"The run ended as **{run_status}**, not SUCCEEDED — this usually means the "
                               "Apify monthly free-tier limit was hit or the actor errored. Check the Apify console.")
                else:
                    st.warning("The run succeeded but matched nobody. This location + role combination may be "
                               "too narrow (e.g. a small town with an abbreviated state). Try a larger metro "
                               "(e.g. **Boston** instead of Chelmsford, MA) or a broader role, then refine.")
                st.stop()
            # optional: keep only people flagged "Open to Work"
            if otw_only:
                found_n = len(recs)
                recs = [r for r in recs if r.get('Open To Work') == 'Yes']
                if not recs:
                    st.warning(f"None of the {found_n} {role_name} found are flagged \"Open to Work\". "
                               "Uncheck that filter, raise the profile count, or widen the location.")
                    st.stop()
            # highest Fit Score first (ties broken by company)
            recs.sort(key=lambda r: (-r['Fit Score'], r['Current Company']))
            # backfill target: keep the best `size` matching results (we over-pulled above)
            recs = recs[:size]

            # ---- dedup against the running master list ----
            master = load_master()
            seen = {_norm_url(m.get('LinkedIn URL', '')) for m in master}
            new_recs, dupes = [], 0
            for r in recs:
                k = _norm_url(r['LinkedIn URL'])
                if k and k not in seen:
                    seen.add(k); new_recs.append(r)
                else:
                    dupes += 1

            # ---- persist: new reps -> master, this run -> history ----
            today = datetime.date.today().isoformat()
            now = datetime.datetime.now().isoformat(timespec='minutes').replace('T', ' ')
            save_master([{'Date Added': today, 'Sourced Role': role_name,
                          'Search Location': location_label, **r} for r in new_recs])
            search_rec = {'Timestamp': now, 'Role': role_name, 'Location': location_label, 'Requested': size,
                          'Found': len(recs), 'New': len(new_recs), 'Dupes': dupes}
            save_search(search_rec)
            summary = {**search_rec, 'MasterTotal': len(master) + len(new_recs)}

            # flag which found people are new vs. already in the master (display only)
            new_urls = {_norm_url(r['LinkedIn URL']) for r in new_recs}
            for r in recs:
                r['New?'] = '🆕 New' if _norm_url(r.get('LinkedIn URL', '')) in new_urls else '• In master'
            disp_cols = ['Fit Score', 'New?'] + [c for c in COLUMNS if c != 'Fit Score']
            st.session_state['last_run'] = {
                'recs': recs, 'disp_cols': disp_cols, 'summary': summary,
                'role': role_name, 'location': location_label, 'remote': remote_only, 'otw_only': otw_only,
                'size': size, 'new': len(new_recs), 'dupes': dupes,
                'n_comp': sum(1 for r in recs if r.get('Competitor')),
                'master_total': summary['MasterTotal'], 'today': today,
            }
        except Exception as e:
            st.error(f"Something went wrong talking to Apify: {e}")

    # ---- render the latest run (kept in session so shortlist selection survives reruns) ----
    _run = st.session_state.get('last_run')
    if _run:
        recs = _run['recs']; disp_cols = _run['disp_cols']
        where = "remotely (nationwide)" if _run['remote'] else f"in {_run['location']}"
        otw_tag = " &nbsp;·&nbsp; open to work" if _run['otw_only'] else ""
        st.markdown(
            f'<div class="ae-result">✓ Found {len(recs)} {_run["role"]} {where}{otw_tag}'
            f' &nbsp;·&nbsp; <b>{_run["new"]} new</b> added to master'
            f' &nbsp;·&nbsp; {_run["dupes"]} already on file'
            f' &nbsp;·&nbsp; {_run["n_comp"]} from competitors</div>', unsafe_allow_html=True)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Profiles found", len(recs))
        m2.metric("New", _run['new'])
        m3.metric("Duplicates", _run['dupes'])
        m4.metric("Master total", _run['master_total'])

        st.caption("Tick ✅ Select for anyone you want, then add them to your shortlist.")
        edit_df = pd.DataFrame(recs, columns=disp_cols)
        edit_df.insert(0, "✅ Select", False)
        edited = st.data_editor(
            edit_df, use_container_width=True, hide_index=True, key="results_editor",
            column_config={"✅ Select": st.column_config.CheckboxColumn("✅ Select")},
            disabled=[c for c in edit_df.columns if c != "✅ Select"])
        a1, a2 = st.columns([1, 2])
        if a1.button("➕ Add selected to Shortlist", use_container_width=True):
            picked = [recs[i] for i in edited.index[edited["✅ Select"] == True].tolist()]
            rows = [{"Added": _run['today'], "Sourced Role": _run['role'],
                     "Fit Score": r.get("Fit Score", ""), "First Name": r.get("First Name", ""),
                     "Last Name": r.get("Last Name", ""), "Email": r.get("Email", ""),
                     "Current Title": r.get("Current Title", ""), "Current Company": r.get("Current Company", ""),
                     "Location": r.get("Location", ""), "LinkedIn URL": r.get("LinkedIn URL", ""),
                     "Status": ""} for r in picked]
            added = add_to_shortlist(rows)
            if added:
                st.success(f"Added {added} to your shortlist — open the ⭐ Shortlist tab to email them.")
            else:
                st.info("Nothing new added (either none were selected or they're already shortlisted).")
        if len(recs) < _run['size']:
            st.caption(f"ℹ️ Only {len(recs)} matched — the {_run['role']}"
                       f"{' open-to-work' if _run['otw_only'] else ''} pool for {_run['location']} is smaller "
                       f"than the {_run['size']} requested (backfill already over-pulled to try to reach it).")
        _loc_slug = re.sub(r'[^A-Za-z0-9]+', '_', str(_run['location'])).strip('_') or 'search'
        st.download_button("⬇ Download Excel",
                           build_excel_bytes(recs, cols=disp_cols, summary=_run['summary']),
                           file_name=f"sales_reps_{_loc_slug}_{_run['today']}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.markdown(
            '<div class="ae-empty">'
            '<div class="big">🔍</div>'
            '<div class="t">Your sourced reps will appear here</div>'
            '<div class="s">Pick a role, location and size above, then run a search to build your list.</div>'
            '</div>', unsafe_allow_html=True)

# ---------------------------- Previous Searches ----------------------------
with tab_history:
    st.markdown('<p class="ae-label">🕘 Previous Searches</p>', unsafe_allow_html=True)
    st.markdown('<p class="ae-help">Every run, newest first — with new vs. duplicate counts.</p>',
                unsafe_allow_html=True)
    searches = load_searches()
    if not searches:
        st.markdown(
            '<div class="ae-empty"><div class="big">🕘</div>'
            '<div class="t">No searches yet</div>'
            '<div class="s">Run a search and it will be logged here automatically.</div></div>',
            unsafe_allow_html=True)
    else:
        hdf = pd.DataFrame(searches)
        for c in SEARCH_HEADERS:
            if c not in hdf.columns:
                hdf[c] = ''
        hdf = hdf[SEARCH_HEADERS].iloc[::-1].reset_index(drop=True)  # newest first
        total_new = pd.to_numeric(hdf['New'], errors='coerce').fillna(0).astype(int).sum()
        c1, c2, c3 = st.columns(3)
        c1.metric("Total searches", len(hdf))
        c2.metric("Profiles added (all time)", int(total_new))
        c3.metric("Master list size", len(load_master()))
        st.dataframe(hdf, use_container_width=True, hide_index=True)

# ------------------------------- Master List -------------------------------
with tab_master:
    st.markdown('<p class="ae-label">📒 Master List</p>', unsafe_allow_html=True)
    st.markdown('<p class="ae-help">Every unique rep sourced so far, deduped by LinkedIn URL.</p>',
                unsafe_allow_html=True)
    master = load_master()
    if not master:
        st.markdown(
            '<div class="ae-empty"><div class="big">📒</div>'
            '<div class="t">Master list is empty</div>'
            '<div class="s">New reps from each search land here automatically.</div></div>',
            unsafe_allow_html=True)
    else:
        mdf = pd.DataFrame(master)
        # Search Location stays in the Google Sheet as metadata but isn't displayed.
        order = [c for c in MASTER_HEADERS if c in mdf.columns and c != 'Search Location']
        mdf = mdf[order]
        if 'Fit Score' in mdf.columns:  # highest Fit Score first
            mdf = (mdf.assign(_fs=pd.to_numeric(mdf['Fit Score'], errors='coerce').fillna(-1))
                      .sort_values('_fs', ascending=False).drop(columns='_fs').reset_index(drop=True))

        # ---- filters ----
        with st.container(border=True):
            roles_present = sorted(x for x in mdf.get('Sourced Role', pd.Series(dtype=str)).dropna().unique() if x)
            f1, f2 = st.columns([2, 1])
            sel_roles = f1.multiselect("Filter by role", roles_present, default=[],
                                       placeholder="All roles")
            min_fit = f2.slider("Min Fit Score", 0, 100, 0, 5)
            g1, g2, g3 = st.columns([2, 1, 1])
            q = g1.text_input("Search name / company / title / location", placeholder="type to filter…")
            only_otw = g2.checkbox("Open to Work only")
            only_comp = g3.checkbox("From competitors")

        fdf = mdf
        if sel_roles and 'Sourced Role' in fdf:
            fdf = fdf[fdf['Sourced Role'].isin(sel_roles)]
        if 'Fit Score' in fdf and min_fit > 0:
            fdf = fdf[pd.to_numeric(fdf['Fit Score'], errors='coerce').fillna(0) >= min_fit]
        if only_otw and 'Open To Work' in fdf:
            fdf = fdf[fdf['Open To Work'] == 'Yes']
        if only_comp and 'Competitor' in fdf:
            fdf = fdf[fdf['Competitor'].astype(str).str.strip() != '']
        if q:
            ql = q.lower()
            search_cols = [c for c in ['First Name', 'Last Name', 'Current Company',
                                       'Current Title', 'Location'] if c in fdf]
            mask = pd.Series(False, index=fdf.index)
            for c in search_cols:
                mask |= fdf[c].astype(str).str.lower().str.contains(ql, na=False, regex=False)
            fdf = fdf[mask]
        fdf = fdf.reset_index(drop=True)

        n_filtered = " (filtered)" if len(fdf) != len(mdf) else ""
        st.markdown(f'<div class="ae-result">📒 Showing {len(fdf)} of {len(mdf)} unique reps{n_filtered}</div>',
                    unsafe_allow_html=True)
        st.dataframe(fdf, use_container_width=True, hide_index=True)
        stamp = datetime.date.today().isoformat()
        st.download_button("⬇ Download master (Excel)",
                           build_excel_bytes(fdf.to_dict('records'), cols=order),
                           file_name=f"sales_reps_master_{stamp}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_master")

# ------------------------------- Shortlist -------------------------------
with tab_shortlist:
    st.markdown('<p class="ae-label">⭐ Shortlist</p>', unsafe_allow_html=True)
    st.markdown('<p class="ae-help">People you picked from search results. Compose, preview, and export outreach here.</p>',
                unsafe_allow_html=True)
    sl = load_shortlist()
    if not sl:
        st.markdown(
            '<div class="ae-empty"><div class="big">⭐</div>'
            '<div class="t">Your shortlist is empty</div>'
            '<div class="s">On the New Search tab, tick people and click "Add selected to Shortlist".</div></div>',
            unsafe_allow_html=True)
    else:
        sldf = pd.DataFrame(sl)
        so = [c for c in SHORTLIST_HEADERS if c in sldf.columns]
        sldf = sldf[so]
        n_email = sum(1 for r in sl if str(r.get('Email', '')).strip())
        st.markdown(f'<div class="ae-result">⭐ {len(sl)} on the shortlist · {n_email} have an email on file</div>',
                    unsafe_allow_html=True)
        st.dataframe(sldf, use_container_width=True, hide_index=True)

        with st.expander("Remove people from the shortlist"):
            labels = [f"{r.get('First Name','')} {r.get('Last Name','')} — {r.get('Current Company','')}".strip()
                      for r in sl]
            rem = st.multiselect("Select people to remove", options=list(range(len(sl))),
                                 format_func=lambda i: labels[i])
            if st.button("Remove selected") and rem:
                keep = [r for i, r in enumerate(sl) if i not in set(rem)]
                replace_shortlist(keep)
                st.rerun()

        st.divider()
        st.markdown("#### ✉️ Outreach email")
        st.info("Write your message below, then click **Open in Outlook** next to a person — a compose window "
                "opens in your own Outlook, pre-filled and ready. You review and hit **Send**. Emails come from "
                "your account; nothing is sent automatically.")
        subj = st.text_input("Subject", value=st.session_state.get('email_subject', DEFAULT_SUBJECT),
                             key='email_subject')
        body = st.text_area("Body", value=st.session_state.get('email_body', DEFAULT_BODY),
                            height=280, key='email_body')
        st.caption("Personalization placeholders: `{first_name}` `{last_name}` `{title}` `{company}` `{role}`")

        recipients = [r for r in sl if str(r.get('Email', '')).strip()]
        no_email = len(sl) - len(recipients)
        if recipients:
            names = [f"{r.get('First Name','')} {r.get('Last Name','')} <{r.get('Email','')}>" for r in recipients]
            pick = st.selectbox("Preview for", options=list(range(len(recipients))),
                                format_func=lambda i: names[i])
            pr = recipients[pick]
            st.markdown(f"**To:** {pr.get('Email','')}")
            st.markdown(f"**Subject:** {fill_template(subj, pr)}")
            st.code(fill_template(body, pr))

            st.markdown("##### 📨 Send from your Outlook")
            st.caption("Each button opens a pre-filled compose window in your Outlook — review and Send.")
            cols = st.columns(3)
            for i, r in enumerate(recipients):
                url = outlook_compose_url(r.get('Email', ''), fill_template(subj, r), fill_template(body, r))
                label = f"✉️ {r.get('First Name','')} {r.get('Last Name','')}".strip()
                cols[i % 3].link_button(label, url, use_container_width=True)
        if no_email:
            st.caption(f"⚠️ {no_email} shortlisted "
                       f"{'person has' if no_email == 1 else 'people have'} no email on file — they'll be skipped.")

        st.divider()
        st.caption("Prefer to send from another tool? Export the drafts instead:")
        sender = ""
        try:
            sender = st.secrets.get("EMAIL_FROM", "")
        except Exception:
            pass
        d1, d2 = st.columns(2)
        d1.download_button("⬇ Download drafts (.eml, opens in Outlook)",
                           build_drafts_zip(recipients, subj, body, sender),
                           file_name="outreach_drafts.zip", mime="application/zip",
                           disabled=not recipients, use_container_width=True)
        d2.download_button("⬇ Download as CSV (mail merge)",
                           build_drafts_csv(recipients, subj, body),
                           file_name="outreach_drafts.csv", mime="text/csv",
                           disabled=not recipients, use_container_width=True, key="dl_csv")

st.markdown('<div class="ae-footer">Albireo Energy · Internal sourcing tool · Powered by LinkedIn public data via Apify</div>',
            unsafe_allow_html=True)
